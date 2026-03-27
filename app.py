"""
DSC Incident Tracker - Flask + MySQL API (with Auth + Excel Upload)
Install: pip install flask flask-cors mysql-connector-python python-dotenv bcrypt openpyxl PyJWT
Run:     python app.py
Seed users: python app.py --create-users
"""

import os, sys, math, re
from datetime import date, datetime, timedelta, timezone
from functools import wraps

from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
import mysql.connector
from mysql.connector import pooling
from dotenv import load_dotenv

try:
    import bcrypt
except ImportError:
    sys.exit("Missing bcrypt. Run: pip install bcrypt")

try:
    import jwt
except ImportError:
    sys.exit("Missing PyJWT. Run: pip install PyJWT")

try:
    import openpyxl
except ImportError:
    sys.exit("Missing openpyxl. Run: pip install openpyxl")

import secrets as _secrets

load_dotenv()

app = Flask(__name__)
app.url_map.strict_slashes = False
CORS(app, resources={r"/api/*": {"origins": "*"}})

# ── Persistent SECRET_KEY ──────────────────────────────────
# Priority: .env  →  .jwt_secret file  →  generate + save to .jwt_secret
# This means Flask can restart, crash, or lose its .env and the key NEVER changes.
def _get_or_create_secret_key():
    # 1. Honour explicit .env / environment variable if set
    env_key = os.getenv("SECRET_KEY", "").strip()
    if env_key and env_key != "change-this-to-a-random-string-in-production":
        print("[AUTH] SECRET_KEY loaded from environment / .env")
        return env_key

    # 2. Load from persisted key file
    key_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".jwt_secret")
    if os.path.exists(key_file):
        with open(key_file, "r") as f:
            key = f.read().strip()
        if key:
            print(f"[AUTH] SECRET_KEY loaded from {key_file}")
            return key

    # 3. Generate a new random key and persist it so future restarts use the same key
    key = _secrets.token_hex(48)
    with open(key_file, "w") as f:
        f.write(key)
    print(f"[AUTH] Generated new SECRET_KEY and saved to {key_file}")
    return key

SECRET_KEY = _get_or_create_secret_key()
TOKEN_HOURS    = int(os.getenv("TOKEN_HOURS", 8))
UPLOAD_LIMIT   = 5 * 1024 * 1024   # 5 MB

# ── DB Pool ────────────────────────────────────────────────
pool = pooling.MySQLConnectionPool(
    pool_name="dsc_pool", pool_size=5,
    host=os.getenv("DB_HOST", "localhost"),
    port=int(os.getenv("DB_PORT", 3306)),
    user=os.getenv("DB_USER", "root"),
    password=os.getenv("DB_PASSWORD", ""),
    database=os.getenv("DB_NAME", "dsc_incidents"),
    charset="utf8mb4", collation="utf8mb4_unicode_ci",
)

def get_conn():
    return pool.get_connection()

# ── Helpers ────────────────────────────────────────────────
def serialize(row):
    if not row: return row
    for k, v in row.items():
        if isinstance(v, (date, datetime)):
            row[k] = v.isoformat()
    return row

def ok(data=None, message="Success", status=200, **kwargs):
    body = {"success": True, "message": message}
    if data is not None:
        body["data"] = data
    body.update(kwargs)
    return jsonify(body), status

def err(message, code=400):
    return jsonify({"success": False, "message": message}), code

ALLOWED_STATUSES = {"Closed", "Pending", "Open"}
SORT_COLUMNS     = {"id", "incident_date", "customer", "category",
                    "resource", "status", "created_at"}

# ── JWT Auth ───────────────────────────────────────────────
def make_token(user):
    payload = {
        "sub":       str(user["id"]),   # PyJWT 2.x requires sub to be a string
        "username":  user["username"],
        "role":      user["role"],
        "full_name": user["full_name"],
        "can_delete": bool(user.get("can_delete", 0)),
        "exp":       datetime.now(timezone.utc) + timedelta(hours=TOKEN_HOURS),
    }
    token = jwt.encode(payload, SECRET_KEY, algorithm="HS256")
    # PyJWT < 2.0 returns bytes; >= 2.0 returns str
    if isinstance(token, bytes):
        token = token.decode("utf-8")
    return token

def decode_token(token):
    token = token.strip()
    return jwt.decode(token, SECRET_KEY, algorithms=["HS256"])

def get_token_from_request():
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        return auth[7:].strip()
    return request.args.get("token", "").strip()

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        token = get_token_from_request()
        if not token:
            return err("Authentication required", 401)
        try:
            request.user = decode_token(token)
        except jwt.ExpiredSignatureError:
            return err("Session expired — please log in again", 401)
        except Exception as e:
            print(f"[TOKEN ERROR] {type(e).__name__}: {e}")
            return err("Invalid token — please log out and log in again", 401)
        return f(*args, **kwargs)
    return wrapper

def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        token = get_token_from_request()
        if not token:
            return err("Authentication required", 401)
        try:
            request.user = decode_token(token)
        except jwt.ExpiredSignatureError:
            return err("Session expired — please log in again", 401)
        except Exception as e:
            print(f"[TOKEN ERROR] {type(e).__name__}: {e}")
            return err("Invalid token — please log out and log in again", 401)
        if request.user.get("role") != "admin":
            return err("Admin access required", 403)
        return f(*args, **kwargs)
    return wrapper

# ── Seed default users ─────────────────────────────────────
def seed_users():
    conn = get_conn()
    cur  = conn.cursor(dictionary=True)
    defaults = [
        ("admin",  "Admin@123",  "Administrator", "admin"),
        ("viewer", "Viewer@123", "Viewer User",   "user"),
    ]
    for username, password, full_name, role in defaults:
        cur.execute("SELECT id FROM users WHERE username = %s", (username,))
        if not cur.fetchone():
            ph = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
            cur.execute(
                "INSERT INTO users (username, password_hash, full_name, role) VALUES (%s,%s,%s,%s)",
                (username, ph, full_name, role)
            )
            print(f"  Created {role}: {username} / {password}")
        else:
            print(f"  User '{username}' already exists — skipped")
    conn.commit()
    cur.close(); conn.close()

# ── resolve or insert FK helpers ───────────────────────────
def resolve_fk(cur, table, name):
    cur.execute(f"SELECT id FROM {table} WHERE name = %s", (name,))
    row = cur.fetchone()
    return row["id"] if row else None

def resolve_or_insert(cur, conn, table, name):
    fk = resolve_fk(cur, table, name)
    if not fk:
        cur.execute(f"INSERT INTO {table} (name) VALUES (%s)", (name,))
        conn.commit()
        fk = cur.lastrowid
    return fk

# ── Excel column map ───────────────────────────────────────
EXCEL_COL_MAP = {
    "s.no": None,
    "customer": "customer",
    "ca": "ca",
    "incident id": "incident_ref",
    "category": "category",
    "date": "incident_date",
    "issue reported": "issue_reported",
    "action taken": "action_taken",
    "resource": "resource",
    "status": "status",
    "no of days taken": "days_taken",
    "hours spent": "hours_spent",
    "device model": "device_model",
    "firmware version": "firmware_version",
    "remarks": "remarks",
}

def parse_excel_date(val):
    if val is None: return None
    if isinstance(val, (datetime, date)): return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

# ══════════════════════════════════════════════════════════
#  ROUTES
# ══════════════════════════════════════════════════════════

# ── Serve frontend ─────────────────────────────────────────
@app.route("/", methods=["GET"])
def index():
    return send_from_directory(
        os.path.dirname(os.path.abspath(__file__)),
        "incident_report_db.html"
    )

# ── Health ─────────────────────────────────────────────────
@app.route("/api/health", methods=["GET"])
def health():
    try:
        conn = get_conn(); conn.ping(reconnect=True); conn.close()
        return ok(message="Database connection healthy")
    except Exception as e:
        return err(f"DB unreachable: {e}", 503)

# ── Auth: Login ────────────────────────────────────────────
@app.route("/api/auth/login", methods=["POST"])
def login():
    if not request.is_json:
        return err("Content-Type must be application/json")
    body = request.get_json()
    username = (body.get("username") or "").strip()
    password = (body.get("password") or "")
    if not username or not password:
        return err("Username and password required")

    conn = get_conn()
    cur  = conn.cursor(dictionary=True)
    try:
        cur.execute(
            "SELECT id, username, password_hash, full_name, role, is_active, can_delete FROM users WHERE username = %s",
            (username,)
        )
        user = cur.fetchone()
        if not user or not bcrypt.checkpw(password.encode(), user["password_hash"].encode()):
            return err("Invalid username or password", 401)
        if not user["is_active"]:
            return err("Account is disabled", 403)
        cur.execute("UPDATE users SET last_login = NOW() WHERE id = %s", (user["id"],))
        conn.commit()
        token = make_token(user)
        return ok({
            "token":      token,
            "username":   user["username"],
            "full_name":  user["full_name"],
            "role":       user["role"],
            "can_delete": bool(user.get("can_delete", 0)),
            "expires_in": TOKEN_HOURS * 3600,
        }, message="Login successful")
    finally:
        cur.close(); conn.close()

# ── Auth: Me ───────────────────────────────────────────────
@app.route("/api/auth/me", methods=["GET"])
@login_required
def me():
    return ok({
        "username":  request.user["username"],
        "full_name": request.user["full_name"],
        "role":      request.user["role"],
    })

# ── Auth: Refresh token ───────────────────────────────────
@app.route("/api/auth/refresh", methods=["POST"])
@login_required
def refresh_token():
    """Re-issue a fresh token for the currently authenticated user."""
    conn = get_conn()
    cur  = conn.cursor(dictionary=True)
    try:
        cur.execute(
            "SELECT id, username, full_name, role, is_active, can_delete FROM users WHERE id = %s",
            (int(request.user["sub"]),)
        )
        user = cur.fetchone()
        if not user or not user["is_active"]:
            return err("Account not found or disabled", 403)
        token = make_token(user)
        return ok({
            "token":      token,
            "expires_in": TOKEN_HOURS * 3600,
        }, message="Token refreshed")
    finally:
        cur.close(); conn.close()

# ── Auth: Change password ──────────────────────────────────
@app.route("/api/auth/change-password", methods=["POST"])
@login_required
def change_password():
    if not request.is_json:
        return err("Content-Type must be application/json")
    body = request.get_json()
    old_pw = (body.get("old_password") or "")
    new_pw = (body.get("new_password") or "")
    if not old_pw or not new_pw:
        return err("old_password and new_password required")
    if len(new_pw) < 6:
        return err("New password must be at least 6 characters")

    conn = get_conn()
    cur  = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT password_hash FROM users WHERE id = %s", (int(request.user["sub"]),))
        user = cur.fetchone()
        if not bcrypt.checkpw(old_pw.encode(), user["password_hash"].encode()):
            return err("Current password is incorrect", 401)
        new_hash = bcrypt.hashpw(new_pw.encode(), bcrypt.gensalt()).decode()
        cur.execute("UPDATE users SET password_hash = %s WHERE id = %s", (new_hash, int(request.user["sub"])))
        conn.commit()
        return ok(message="Password changed successfully")
    finally:
        cur.close(); conn.close()

# ── Admin: List users ──────────────────────────────────────
@app.route("/api/admin/users", methods=["GET", "POST"])
@admin_required
def admin_users():
    if request.method == "GET":
        conn = get_conn(); cur = conn.cursor(dictionary=True)
        try:
            cur.execute("SELECT id, username, full_name, email, role, is_active, last_login, created_at FROM users ORDER BY created_at DESC")
            return ok([serialize(r) for r in cur.fetchall()])
        finally:
            cur.close(); conn.close()

    # POST — create user
    if not request.is_json:
        return err("Content-Type must be application/json")
    body = request.get_json()
    required = ["username", "password", "full_name", "role"]
    missing = [f for f in required if not body.get(f)]
    if missing:
        return err(f"Missing: {', '.join(missing)}")
    if body["role"] not in ("admin", "user"):
        return err("role must be admin or user")
    if len(body["password"]) < 6:
        return err("Password must be at least 6 characters")

    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id FROM users WHERE username = %s", (body["username"],))
        if cur.fetchone():
            return err(f"Username '{body['username']}' already exists")
        ph = bcrypt.hashpw(body["password"].encode(), bcrypt.gensalt()).decode()
        cur.execute(
            "INSERT INTO users (username, password_hash, full_name, email, role) VALUES (%s,%s,%s,%s,%s)",
            (body["username"], ph, body["full_name"], body.get("email"), body["role"])
        )
        conn.commit()
        return ok({"id": cur.lastrowid, "username": body["username"], "role": body["role"]},
                  message="User created", status=201)
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"Database error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

@app.route("/api/admin/users/<int:user_id>", methods=["PUT", "DELETE"])
@admin_required
def admin_user_item(user_id):
    if request.method == "DELETE":
        if user_id == int(request.user["sub"]):
            return err("Cannot delete your own account")
        conn = get_conn(); cur = conn.cursor()
        try:
            cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
            if cur.rowcount == 0: return err("User not found", 404)
            conn.commit()
            return ok(message="User deleted")
        finally:
            cur.close(); conn.close()

    # PUT — update user
    if not request.is_json:
        return err("Content-Type must be application/json")
    body = request.get_json()
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id FROM users WHERE id = %s", (user_id,))
        if not cur.fetchone(): return err("User not found", 404)
        sets, params = [], []
        if "full_name" in body: sets.append("full_name = %s"); params.append(body["full_name"])
        if "email"     in body: sets.append("email = %s");     params.append(body["email"])
        if "role"      in body:
            if body["role"] not in ("admin", "user"): return err("role must be admin or user")
            sets.append("role = %s"); params.append(body["role"])
        if "is_active" in body: sets.append("is_active = %s"); params.append(int(body["is_active"]))
        if "password"  in body:
            if len(body["password"]) < 6: return err("Password must be at least 6 characters")
            ph = bcrypt.hashpw(body["password"].encode(), bcrypt.gensalt()).decode()
            sets.append("password_hash = %s"); params.append(ph)
        if not sets: return err("No fields to update")
        params.append(user_id)
        cur.execute(f"UPDATE users SET {', '.join(sets)} WHERE id = %s", params)
        conn.commit()
        return ok(message="User updated")
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"Database error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

# ── Admin: Excel Upload ────────────────────────────────────
@app.route("/api/admin/upload-excel", methods=["POST"])
@admin_required
def upload_excel():
    if "file" not in request.files:
        return err("No file uploaded. Field name must be 'file'")
    f = request.files["file"]
    if not f.filename:
        return err("No file selected")
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return err("Only .xlsx or .xls files are accepted")

    content = f.read()
    if len(content) > UPLOAD_LIMIT:
        return err(f"File too large (max {UPLOAD_LIMIT//1024//1024} MB)")

    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return err(f"Cannot read Excel file: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return err("Excel file has no data rows")

    # Map header row to field names
    header_raw = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    col_map = {}
    for idx, h in enumerate(header_raw):
        if h in EXCEL_COL_MAP:
            col_map[idx] = EXCEL_COL_MAP[h]

    if not col_map:
        return err("No recognisable column headers found. Expected: Customer, Category, Date, Issue Reported, Action Taken, Resource, Status…")

    conn = get_conn()
    cur  = conn.cursor(dictionary=True)

    inserted, skipped, errors = 0, 0, []

    for row_num, row in enumerate(rows[1:], start=2):
        record = {}
        for idx, field in col_map.items():
            if field is None: continue
            val = row[idx] if idx < len(row) else None
            if val is not None:
                record[field] = str(val).strip() if not isinstance(val, (date, datetime)) else val

        # Skip completely empty rows
        if not any(record.values()):
            skipped += 1; continue

        # Required fields
        missing = [f for f in ["customer","category","incident_date","issue_reported","resource"] if not record.get(f)]
        if missing:
            errors.append(f"Row {row_num}: missing {', '.join(missing)}")
            skipped += 1; continue

        # Parse date
        record["incident_date"] = parse_excel_date(record.get("incident_date"))
        if not record["incident_date"]:
            errors.append(f"Row {row_num}: invalid date '{record.get('incident_date')}'")
            skipped += 1; continue

        # Status default
        status = record.get("status", "Open")
        if status not in ALLOWED_STATUSES:
            status = "Open"

        # Resolve FKs
        try:
            c_id   = resolve_fk(cur, "customers",  record["customer"])
            cat_id = resolve_fk(cur, "categories", record["category"])
            if not c_id:
                cur.execute("INSERT IGNORE INTO customers (name) VALUES (%s)", (record["customer"],))
                conn.commit()
                c_id = resolve_fk(cur, "customers", record["customer"])
            if not cat_id:
                cur.execute("INSERT IGNORE INTO categories (name) VALUES (%s)", (record["category"],))
                conn.commit()
                cat_id = resolve_fk(cur, "categories", record["category"])
            r_id = resolve_or_insert(cur, conn, "resources", record["resource"])

            cur.execute("""
                INSERT INTO incidents
                  (customer_id, incident_ref, ca, category_id, incident_date,
                   issue_reported, action_taken, resource_id, status,
                   days_taken, hours_spent, device_model, firmware_version, remarks)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                c_id, record.get("incident_ref"), record.get("ca"), cat_id, record["incident_date"],
                record.get("issue_reported",""), record.get("action_taken"),
                r_id, status,
                record.get("days_taken"), record.get("hours_spent"),
                record.get("device_model"), record.get("firmware_version"),
                record.get("remarks")
            ))
            conn.commit()
            inserted += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {e}")
            skipped += 1

    cur.close(); conn.close()

    return ok({
        "inserted": inserted,
        "skipped":  skipped,
        "errors":   errors[:20],  # cap error list
    }, message=f"Upload complete: {inserted} inserted, {skipped} skipped")

# ── Lookups ────────────────────────────────────────────────
@app.route("/api/lookups", methods=["GET"])
@login_required
def get_lookups():
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, name FROM customers  ORDER BY name")
        customers = cur.fetchall()
        cur.execute("SELECT id, name FROM categories ORDER BY name")
        categories = cur.fetchall()
        cur.execute("SELECT id, name FROM resources  ORDER BY name")
        resources = cur.fetchall()
        return ok({"customers": customers, "categories": categories, "resources": resources})
    finally:
        cur.close(); conn.close()

# ── Stats ──────────────────────────────────────────────────
@app.route("/api/stats", methods=["GET"])
@login_required
def get_stats():
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT COUNT(*) AS total,
              SUM(status='Closed') AS closed, SUM(status='Pending') AS pending,
              SUM(status='Open') AS open_count,
              SUM(category='Breakfix') AS breakfix,
              SUM(category='Devlopment Team') AS dev_team,
              SUM(category='Integration') AS integration,
              ROUND(SUM(status='Closed')/COUNT(*)*100,1) AS close_rate,
              ROUND(SUM(days_taken='Same day closed')/COUNT(*)*100,1) AS same_day_rate
            FROM v_incidents
        """)
        summary = cur.fetchone()
        cur.execute("SELECT customer, COUNT(*) AS cnt FROM v_incidents GROUP BY customer ORDER BY cnt DESC")
        by_customer = cur.fetchall()
        cur.execute("SELECT resource, COUNT(*) AS cnt FROM v_incidents GROUP BY resource ORDER BY cnt DESC")
        by_resource = cur.fetchall()
        cur.execute("SELECT DATE_FORMAT(incident_date,'%Y-%m') AS month, COUNT(*) AS cnt FROM v_incidents GROUP BY month ORDER BY month")
        by_month = cur.fetchall()
        return ok({"summary": summary, "by_customer": by_customer,
                   "by_resource": by_resource, "by_month": by_month})
    finally:
        cur.close(); conn.close()

# ── Excel Export (login_required — both viewer & admin) ───
@app.route("/api/incidents/export", methods=["GET"])
@login_required
def export_incidents():
    import io
    from flask import send_file

    status    = request.args.get("status",   "")
    category  = request.args.get("category", "")
    search    = request.args.get("search",   "").strip()
    date_from = request.args.get("date_from","")
    date_to   = request.args.get("date_to",  "")

    where, params = [], []
    if status and status in ALLOWED_STATUSES:
        where.append("status = %s"); params.append(status)
    if category:
        where.append("category = %s"); params.append(category)
    if search:
        where.append("(issue_reported LIKE %s OR action_taken LIKE %s OR remarks LIKE %s)")
        like = f"%{search}%"; params.extend([like, like, like])
    if date_from:
        where.append("incident_date >= %s"); params.append(date_from)
    if date_to:
        where.append("incident_date <= %s"); params.append(date_to)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"SELECT * FROM v_incidents {where_sql} ORDER BY incident_date DESC, created_at DESC",
            params
        )
        rows = cur.fetchall()
    finally:
        cur.close(); conn.close()

    # ── Build workbook ──────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incidents"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Header style
    header_fill = PatternFill("solid", fgColor="1E40AF")   # deep blue
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="D1D5DB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    COLUMNS = [
        ("S.No",             12),
        ("Customer",         20),
        ("Incident ID",      18),
        ("CA",               18),
        ("Category",         18),
        ("Date",             14),
        ("Issue Reported",   40),
        ("Action Taken",     40),
        ("Resource",         16),
        ("Status",           12),
        ("Days Taken",       18),
        ("Hours Spent",      14),
        ("Device Model",     16),
        ("Firmware Version", 18),
        ("Remarks",          30),
        ("Created At",       20),
    ]

    # Write headers
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill       = header_fill
        cell.font       = header_font
        cell.alignment  = center
        cell.border     = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 30

    # Status fill colours
    STATUS_FILLS = {
        "Closed":  PatternFill("solid", fgColor="D1FAE5"),
        "Pending": PatternFill("solid", fgColor="FEF3C7"),
        "Open":    PatternFill("solid", fgColor="FEE2E2"),
    }
    STATUS_FONTS = {
        "Closed":  Font(color="065F46", bold=True),
        "Pending": Font(color="92400E", bold=True),
        "Open":    Font(color="991B1B", bold=True),
    }

    # Write data rows
    for r_idx, row in enumerate(rows, start=2):
        status_val = row.get("status", "")
        row_fill   = STATUS_FILLS.get(status_val)
        values = [
            r_idx - 1,
            row.get("customer"),
            row.get("incident_ref"),
            row.get("ca"),
            row.get("category"),
            row.get("incident_date").isoformat() if isinstance(row.get("incident_date"), date) else row.get("incident_date"),
            row.get("issue_reported"),
            row.get("action_taken"),
            row.get("resource"),
            status_val,
            row.get("days_taken"),
            row.get("hours_spent"),
            row.get("device_model"),
            row.get("firmware_version"),
            row.get("remarks"),
            row.get("created_at").isoformat() if isinstance(row.get("created_at"), datetime) else str(row.get("created_at") or ""),
        ]
        for c_idx, val in enumerate(values, start=1):
            cell            = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border     = border
            cell.alignment  = center if c_idx in (1, 5, 6, 9, 10, 11, 12, 13, 14) else left
            if c_idx == 10 and row_fill:          # status column
                cell.fill = row_fill
                cell.font = STATUS_FONTS.get(status_val, Font())
            elif r_idx % 2 == 0:                 # zebra stripe
                cell.fill = PatternFill("solid", fgColor="F8FAFC")
        ws.row_dimensions[r_idx].height = 20

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Output
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import date as _date
    fname = f"DSC_Incidents_{_date.today().isoformat()}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )

# ── Incidents collection ───────────────────────────────────
@app.route("/api/incidents", methods=["GET", "POST"])
@login_required
def incidents_collection():
    if request.method == "GET":
        return _list_incidents()
    return _create_incident()

def _list_incidents():
    page     = max(1, int(request.args.get("page", 1)))
    per_page = min(200, max(1, int(request.args.get("per_page", 50))))
    offset   = (page - 1) * per_page
    status   = request.args.get("status", "")
    category = request.args.get("category", "")
    search   = request.args.get("search", "").strip()
    date_from= request.args.get("date_from", "")
    date_to  = request.args.get("date_to", "")
    sort     = request.args.get("sort", "incident_date")
    order    = request.args.get("order", "desc").upper()
    if sort  not in SORT_COLUMNS:  sort  = "incident_date"
    if order not in ("ASC","DESC"): order = "DESC"

    where, params = [], []
    if status and status in ALLOWED_STATUSES:
        where.append("status = %s"); params.append(status)
    if category:
        where.append("category = %s"); params.append(category)
    if search:
        where.append("(issue_reported LIKE %s OR action_taken LIKE %s OR remarks LIKE %s)")
        like = f"%{search}%"; params.extend([like, like, like])
    if date_from:
        where.append("incident_date >= %s"); params.append(date_from)
    if date_to:
        where.append("incident_date <= %s"); params.append(date_to)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute(f"SELECT COUNT(*) AS total FROM v_incidents {where_sql}", params)
        total = cur.fetchone()["total"]
        cur.execute(
            f"SELECT * FROM v_incidents {where_sql} ORDER BY {sort} {order} LIMIT %s OFFSET %s",
            params + [per_page, offset]
        )
        rows = [serialize(r) for r in cur.fetchall()]
        return ok(rows, total=total, page=page, per_page=per_page,
                  total_pages=math.ceil(total / per_page) if total else 1)
    finally:
        cur.close(); conn.close()

def _create_incident():
    if not request.is_json: return err("Content-Type must be application/json")
    body = request.get_json()
    missing = [f for f in ["customer","category","incident_date","issue_reported","resource"] if not body.get(f)]
    if missing: return err(f"Missing required fields: {', '.join(missing)}")
    status = body.get("status", "Open")
    if status not in ALLOWED_STATUSES: return err(f"status must be one of: {', '.join(ALLOWED_STATUSES)}")
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        c_id   = resolve_fk(cur, "customers",  body["customer"])
        if not c_id: return err(f"Unknown customer: {body['customer']}")
        cat_id = resolve_fk(cur, "categories", body["category"])
        if not cat_id: return err(f"Unknown category: {body['category']}")
        r_id   = resolve_or_insert(cur, conn, "resources", body["resource"].strip())
        cur.execute("""
            INSERT INTO incidents
              (customer_id, incident_ref, ca, category_id, incident_date,
               issue_reported, action_taken, resource_id, status,
               days_taken, hours_spent, device_model, firmware_version, remarks)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (c_id, body.get("incident_ref"), body.get("ca"), cat_id, body["incident_date"],
              body["issue_reported"], body.get("action_taken"), r_id, status,
              body.get("days_taken"), body.get("hours_spent"), body.get("device_model"),
              body.get("firmware_version"), body.get("remarks")))
        conn.commit()
        new_id = cur.lastrowid
        cur.execute("SELECT * FROM v_incidents WHERE id = %s", (new_id,))
        return ok(serialize(cur.fetchone()), message="Incident created", status=201)
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"Database error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

# ── Incidents item ─────────────────────────────────────────
@app.route("/api/incidents/<int:incident_id>", methods=["GET", "PUT", "DELETE"])
@login_required
def incidents_item(incident_id):
    if request.method == "GET":    return _get_incident(incident_id)
    if request.method == "PUT":    return _update_incident(incident_id)
    return _delete_incident(incident_id)

def _get_incident(incident_id):
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT * FROM v_incidents WHERE id = %s", (incident_id,))
        row = cur.fetchone()
        if not row: return err("Incident not found", 404)
        return ok(serialize(row))
    finally:
        cur.close(); conn.close()

def _update_incident(incident_id):
    if request.user.get("role") != "admin":
        return err("Only admins can update incidents", 403)
    if not request.is_json: return err("Content-Type must be application/json")
    body = request.get_json()
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id FROM incidents WHERE id = %s", (incident_id,))
        if not cur.fetchone(): return err("Incident not found", 404)
        sets, params = [], []
        if "customer" in body:
            fk = resolve_fk(cur, "customers", body["customer"])
            if not fk: return err(f"Unknown customer: {body['customer']}")
            sets.append("customer_id = %s"); params.append(fk)
        if "category" in body:
            fk = resolve_fk(cur, "categories", body["category"])
            if not fk: return err(f"Unknown category: {body['category']}")
            sets.append("category_id = %s"); params.append(fk)
        if "resource" in body:
            fk = resolve_or_insert(cur, conn, "resources", body["resource"].strip())
            sets.append("resource_id = %s"); params.append(fk)
        for f in ["incident_ref","ca","incident_date","issue_reported","action_taken",
                  "status","days_taken","hours_spent","device_model","firmware_version","remarks"]:
            if f in body:
                if f == "status" and body[f] not in ALLOWED_STATUSES:
                    return err(f"status must be one of: {', '.join(ALLOWED_STATUSES)}")
                sets.append(f"{f} = %s"); params.append(body[f])
        if not sets: return err("No fields to update")
        params.append(incident_id)
        cur.execute(f"UPDATE incidents SET {', '.join(sets)} WHERE id = %s", params)
        conn.commit()
        cur.execute("SELECT * FROM v_incidents WHERE id = %s", (incident_id,))
        return ok(serialize(cur.fetchone()), message="Incident updated")
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"Database error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

def _delete_incident(incident_id):
    if not request.user.get("can_delete"):
        return err("You do not have permission to delete incidents", 403)
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("DELETE FROM incidents WHERE id = %s", (incident_id,))
        if cur.rowcount == 0: return err("Incident not found", 404)
        conn.commit(); return ok(message="Incident deleted")
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"Database error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

# ── Delete ALL incidents (senthil only) ───────────────────
@app.route("/api/incidents/delete-all", methods=["DELETE"])
@login_required
def delete_all_incidents():
    if not request.user.get("can_delete"):
        return err("You do not have permission to delete incidents", 403)
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) FROM incidents")
        total = cur.fetchone()[0]
        cur.execute("DELETE FROM incidents")
        conn.commit()
        return ok(message="All incidents deleted", deleted=total)
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"Database error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

# ── Status patch ───────────────────────────────────────────
@app.route("/api/incidents/<int:incident_id>/status", methods=["PATCH"])
@login_required
def patch_status(incident_id):
    if not request.is_json: return err("Content-Type must be application/json")
    body = request.get_json()
    status = body.get("status")
    if status not in ALLOWED_STATUSES: return err(f"status must be one of: {', '.join(ALLOWED_STATUSES)}")
    conn = get_conn(); cur = conn.cursor()
    try:
        cur.execute("UPDATE incidents SET status = %s WHERE id = %s", (status, incident_id))
        if cur.rowcount == 0: return err("Incident not found", 404)
        conn.commit(); return ok(message=f"Status updated to {status}")
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"Database error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

# ══════════════════════════════════════════════════════════
#  INTEGRATION PENDING TRACKER — ROUTES
# ══════════════════════════════════════════════════════════

INT_CA_STATUSES  = {'Pending','Completed','Partially Completed','Working In Progress'}
INT_PROV_TYPES   = {'device','sign','service'}
INT_ISSUE_STATUSES = {
    'Resolved','Un Resolved','Integration Required',
    'On Hold','Devolopment team Is working On'
}

# ── CA List ────────────────────────────────────────────────
@app.route("/api/integration/ca", methods=["GET","POST"])
@login_required
def int_ca():
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        if request.method == "GET":
            status = request.args.get("status","")
            search = request.args.get("search","").strip()
            where, params = [], []
            if status and status in INT_CA_STATUSES:
                where.append("status=%s"); params.append(status)
            if search:
                where.append("name LIKE %s"); params.append(f"%{search}%")
            w = ("WHERE "+" AND ".join(where)) if where else ""
            cur.execute(f"SELECT * FROM int_ca_list {w} ORDER BY id", params)
            return ok([serialize(r) for r in cur.fetchall()])

        # POST — admin only
        if request.user.get("role") != "admin":
            return err("Admin access required", 403)
        if not request.is_json: return err("JSON required")
        body = request.get_json()
        if not body.get("name"): return err("name is required")
        status = body.get("status","Pending")
        if status not in INT_CA_STATUSES: return err(f"Invalid status")
        cur.execute(
            "INSERT INTO int_ca_list (name,status,remarks) VALUES (%s,%s,%s)",
            (body["name"].strip(), status, body.get("remarks"))
        )
        conn.commit()
        cur.execute("SELECT * FROM int_ca_list WHERE id=%s", (cur.lastrowid,))
        return ok(serialize(cur.fetchone()), message="CA entry created", status=201)
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"DB error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

@app.route("/api/integration/ca/<int:rid>", methods=["PUT","DELETE"])
@admin_required
def int_ca_item(rid):
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id FROM int_ca_list WHERE id=%s", (rid,))
        if not cur.fetchone(): return err("Record not found", 404)
        if request.method == "DELETE":
            cur.execute("DELETE FROM int_ca_list WHERE id=%s", (rid,))
            conn.commit(); return ok(message="Deleted")
        if not request.is_json: return err("JSON required")
        body = request.get_json()
        sets, params = [], []
        if "name"    in body: sets.append("name=%s");    params.append(body["name"].strip())
        if "status"  in body:
            if body["status"] not in INT_CA_STATUSES: return err("Invalid status")
            sets.append("status=%s"); params.append(body["status"])
        if "remarks" in body: sets.append("remarks=%s"); params.append(body["remarks"])
        if not sets: return err("Nothing to update")
        params.append(rid)
        cur.execute(f"UPDATE int_ca_list SET {','.join(sets)} WHERE id=%s", params)
        conn.commit()
        cur.execute("SELECT * FROM int_ca_list WHERE id=%s", (rid,))
        return ok(serialize(cur.fetchone()), message="Updated")
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"DB error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

# ── Providers ──────────────────────────────────────────────
@app.route("/api/integration/providers", methods=["GET","POST"])
@login_required
def int_providers():
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        if request.method == "GET":
            ptype = request.args.get("type","")
            where, params = [], []
            if ptype and ptype in INT_PROV_TYPES:
                where.append("provider_type=%s"); params.append(ptype)
            w = ("WHERE "+" AND ".join(where)) if where else ""
            cur.execute(f"SELECT * FROM int_providers {w} ORDER BY provider_type,id", params)
            return ok([serialize(r) for r in cur.fetchall()])

        if request.user.get("role") != "admin":
            return err("Admin access required", 403)
        if not request.is_json: return err("JSON required")
        body = request.get_json()
        if not body.get("name"): return err("name is required")
        ptype  = body.get("provider_type","device")
        status = body.get("status","Pending")
        if ptype  not in INT_PROV_TYPES:    return err("Invalid provider_type")
        if status not in INT_CA_STATUSES:   return err("Invalid status")
        cur.execute(
            "INSERT INTO int_providers (provider_type,name,status,remarks) VALUES (%s,%s,%s,%s)",
            (ptype, body["name"].strip(), status, body.get("remarks"))
        )
        conn.commit()
        cur.execute("SELECT * FROM int_providers WHERE id=%s", (cur.lastrowid,))
        return ok(serialize(cur.fetchone()), message="Provider created", status=201)
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"DB error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

@app.route("/api/integration/providers/<int:rid>", methods=["PUT","DELETE"])
@admin_required
def int_provider_item(rid):
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id FROM int_providers WHERE id=%s", (rid,))
        if not cur.fetchone(): return err("Record not found", 404)
        if request.method == "DELETE":
            cur.execute("DELETE FROM int_providers WHERE id=%s", (rid,))
            conn.commit(); return ok(message="Deleted")
        if not request.is_json: return err("JSON required")
        body = request.get_json()
        sets, params = [], []
        if "name"          in body: sets.append("name=%s");          params.append(body["name"].strip())
        if "provider_type" in body:
            if body["provider_type"] not in INT_PROV_TYPES: return err("Invalid provider_type")
            sets.append("provider_type=%s"); params.append(body["provider_type"])
        if "status"  in body:
            if body["status"] not in INT_CA_STATUSES: return err("Invalid status")
            sets.append("status=%s"); params.append(body["status"])
        if "remarks" in body: sets.append("remarks=%s"); params.append(body["remarks"])
        if not sets: return err("Nothing to update")
        params.append(rid)
        cur.execute(f"UPDATE int_providers SET {','.join(sets)} WHERE id=%s", params)
        conn.commit()
        cur.execute("SELECT * FROM int_providers WHERE id=%s", (rid,))
        return ok(serialize(cur.fetchone()), message="Updated")
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"DB error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

# ── Issues ─────────────────────────────────────────────────
@app.route("/api/integration/issues", methods=["GET","POST"])
@login_required
def int_issues():
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        if request.method == "GET":
            status   = request.args.get("status","")
            search   = request.args.get("search","").strip()
            customer = request.args.get("customer","")
            where, params = [], []
            if status and status in INT_ISSUE_STATUSES:
                where.append("status=%s"); params.append(status)
            if customer:
                where.append("customer=%s"); params.append(customer)
            if search:
                where.append("(issue LIKE %s OR remarks LIKE %s OR customer LIKE %s)")
                like = f"%{search}%"; params.extend([like,like,like])
            w = ("WHERE "+" AND ".join(where)) if where else ""
            cur.execute(f"SELECT * FROM int_issues {w} ORDER BY id", params)
            return ok([serialize(r) for r in cur.fetchall()])

        if request.user.get("role") != "admin":
            return err("Admin access required", 403)
        if not request.is_json: return err("JSON required")
        body = request.get_json()
        if not body.get("customer"): return err("customer is required")
        if not body.get("issue"):    return err("issue is required")
        status = body.get("status","On Hold")
        if status not in INT_ISSUE_STATUSES: return err("Invalid status")
        cur.execute(
            "INSERT INTO int_issues (customer,issue,remarks,status) VALUES (%s,%s,%s,%s)",
            (body["customer"].strip(), body["issue"].strip(), body.get("remarks"), status)
        )
        conn.commit()
        cur.execute("SELECT * FROM int_issues WHERE id=%s", (cur.lastrowid,))
        return ok(serialize(cur.fetchone()), message="Issue created", status=201)
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"DB error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

@app.route("/api/integration/issues/<int:rid>", methods=["PUT","DELETE"])
@admin_required
def int_issue_item(rid):
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id FROM int_issues WHERE id=%s", (rid,))
        if not cur.fetchone(): return err("Record not found", 404)
        if request.method == "DELETE":
            cur.execute("DELETE FROM int_issues WHERE id=%s", (rid,))
            conn.commit(); return ok(message="Deleted")
        if not request.is_json: return err("JSON required")
        body = request.get_json()
        sets, params = [], []
        if "customer" in body: sets.append("customer=%s"); params.append(body["customer"].strip())
        if "issue"    in body: sets.append("issue=%s");    params.append(body["issue"].strip())
        if "remarks"  in body: sets.append("remarks=%s");  params.append(body["remarks"])
        if "status"   in body:
            if body["status"] not in INT_ISSUE_STATUSES: return err("Invalid status")
            sets.append("status=%s"); params.append(body["status"])
        if not sets: return err("Nothing to update")
        params.append(rid)
        cur.execute(f"UPDATE int_issues SET {','.join(sets)} WHERE id=%s", params)
        conn.commit()
        cur.execute("SELECT * FROM int_issues WHERE id=%s", (rid,))
        return ok(serialize(cur.fetchone()), message="Updated")
    except mysql.connector.Error as e:
        conn.rollback(); return err(f"DB error: {e.msg}", 500)
    finally:
        cur.close(); conn.close()

# ── Integration Stats ──────────────────────────────────────
@app.route("/api/integration/stats", methods=["GET"])
@login_required
def int_stats():
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT COUNT(*) AS total, SUM(status='Completed') AS completed, SUM(status='Pending') AS pending, SUM(status='Partially Completed') AS partial FROM int_ca_list")
        ca = cur.fetchone()
        cur.execute("SELECT COUNT(*) AS total, SUM(status='Resolved') AS resolved, SUM(status='Un Resolved') AS unresolved, SUM(status='Integration Required') AS integration_req, SUM(status='On Hold') AS on_hold FROM int_issues")
        issues = cur.fetchone()
        return ok({"ca": ca, "issues": issues})
    finally:
        cur.close(); conn.close()

# ── Integration Excel Upload ───────────────────────────────
@app.route("/api/integration/upload", methods=["POST"])
@admin_required
def int_upload():
    if "file" not in request.files: return err("No file uploaded")
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx",".xls")): return err("Only .xlsx/.xls accepted")
    content = f.read()
    if len(content) > UPLOAD_LIMIT: return err("File too large (max 5MB)")

    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return err(f"Cannot read file: {e}")

    conn = get_conn(); cur = conn.cursor(dictionary=True)
    results = {}

    # ── Sheet 1: CA Integration List ──────────────────────
    if 'CA Integration Pending Status' in wb.sheetnames:
        ws   = wb['CA Integration Pending Status']
        rows = list(ws.iter_rows(values_only=True))
        ins  = 0; skip = 0; errs = []
        for i, row in enumerate(rows[2:], start=3):   # skip 2 header rows
            name = str(row[1]).strip() if row[1] else ""
            stat = str(row[2]).strip() if row[2] else "Pending"
            rem  = str(row[3]).strip() if row[3] else None
            if not name or name in ("None","CA List"): skip += 1; continue
            if stat not in INT_CA_STATUSES: stat = "Pending"
            try:
                cur.execute("SELECT id FROM int_ca_list WHERE name=%s", (name,))
                if cur.fetchone():
                    cur.execute("UPDATE int_ca_list SET status=%s,remarks=%s WHERE name=%s", (stat,rem,name))
                else:
                    cur.execute("INSERT INTO int_ca_list (name,status,remarks) VALUES (%s,%s,%s)", (name,stat,rem))
                conn.commit(); ins += 1
            except Exception as e:
                errs.append(f"Row {i}: {e}"); skip += 1
        results["ca"] = {"upserted": ins, "skipped": skip, "errors": errs[:10]}

    # ── Sheet 2: Issue Fixed/Not Fixed ────────────────────
    if 'Issue Fixed and not Fixed Statu' in wb.sheetnames:
        ws2  = wb['Issue Fixed and not Fixed Statu']
        rows2 = list(ws2.iter_rows(values_only=True))
        ins2 = 0; skip2 = 0; errs2 = []
        STATUS_MAP = {
            'resolved': 'Resolved', 'un resolved': 'Un Resolved',
            'integration required': 'Integration Required',
            'on hold': 'On Hold', 'devolopment': 'Devolopment team Is working On'
        }
        for i, row in enumerate(rows2[1:], start=2):
            customer = str(row[1]).strip() if row[1] else ""
            issue    = str(row[2]).strip() if row[2] else ""
            rem      = str(row[3]).strip() if row[3] else None
            raw_stat = str(row[4]).strip() if row[4] else ""
            if not customer or not issue or customer == "Customer": skip2 += 1; continue
            # Map to valid status
            stat = "On Hold"
            for k, v in STATUS_MAP.items():
                if k in raw_stat.lower(): stat = v; break
            try:
                cur.execute("INSERT INTO int_issues (customer,issue,remarks,status) VALUES (%s,%s,%s,%s)",
                            (customer, issue, rem, stat))
                conn.commit(); ins2 += 1
            except Exception as e:
                errs2.append(f"Row {i}: {e}"); skip2 += 1
        results["issues"] = {"inserted": ins2, "skipped": skip2, "errors": errs2[:10]}

    cur.close(); conn.close()
    return ok(results, message="Upload complete")

# ── Daily Report ───────────────────────────────────────────
@app.route("/api/incidents/daily-report", methods=["GET"])
@login_required
def daily_report():
    """
    Returns today's incidents split into closed/pending.
    'Today' means:
      1. incident_date = report_date  (new calls logged today)
      2. status='Closed' AND days_taken contains today's date string
         (older calls that were finally resolved today — engineer fills
          the closing date in the days_taken field e.g. "27-Mar-26")
    Query param: ?date=YYYY-MM-DD  (defaults to today)
    """
    report_date = request.args.get("date", date.today().isoformat())
    try:
        dt = datetime.strptime(report_date, "%Y-%m-%d")
    except ValueError:
        return err("Invalid date format — use YYYY-MM-DD")

    # Build all the date string formats the engineer might type in days_taken
    # e.g. "27-Mar-26", "27-Mar-2026", "27/03/2026", "27-03-2026"
    months_short = ["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]
    mon_short = months_short[dt.month - 1]
    date_patterns = [
        f"{dt.day:02d}-{mon_short}-{str(dt.year)[2:]}",   # 27-Mar-26
        f"{dt.day:02d}-{mon_short}-{dt.year}",             # 27-Mar-2026
        f"{dt.day}-{mon_short}-{str(dt.year)[2:]}",        # 27-Mar-26 (no zero-pad)
        f"{dt.day}-{mon_short}-{dt.year}",                  # 27-Mar-2026 (no zero-pad)
        f"{dt.day:02d}/{dt.month:02d}/{dt.year}",           # 27/03/2026
        f"{dt.day:02d}-{dt.month:02d}-{dt.year}",          # 27-03-2026
        report_date,                                         # 2026-03-27
    ]

    conn = get_conn()
    cur  = conn.cursor(dictionary=True)
    try:
        # ── 1. Incidents whose incident_date is today ──────────────────
        cur.execute(
            "SELECT * FROM v_incidents WHERE incident_date = %s ORDER BY status DESC, id ASC",
            (report_date,)
        )
        today_rows = [serialize(r) for r in cur.fetchall()]
        today_ids  = {r["id"] for r in today_rows}

        # ── 2. Older incidents closed today via days_taken ─────────────
        # Build a LIKE OR chain for all date formats
        like_clauses = " OR ".join(["days_taken LIKE %s"] * len(date_patterns))
        like_params  = [f"%{p}%" for p in date_patterns]

        cur.execute(
            f"""SELECT * FROM v_incidents
                WHERE status = 'Closed'
                  AND incident_date != %s
                  AND ({like_clauses})
                ORDER BY id ASC""",
            [report_date] + like_params
        )
        earlier_closed = [serialize(r) for r in cur.fetchall()
                          if r["id"] not in today_ids]

        # ── Merge: today rows + earlier-but-closed-today ──────────────
        closed  = [r for r in today_rows if r["status"] == "Closed"] + earlier_closed
        pending = [r for r in today_rows if r["status"] in ("Pending", "Open")]
        all_rows = closed + pending

        # ── Customer-wise pivot (closed + pending) ────────────────────
        summary = {}
        for r in closed:
            c = r["customer"]
            if c not in summary:
                summary[c] = {"Closed": 0, "Pending": 0}
            summary[c]["Closed"] += 1
        for r in pending:
            c = r["customer"]
            if c not in summary:
                summary[c] = {"Closed": 0, "Pending": 0}
            summary[c]["Pending"] += 1

        return ok({
            "report_date":   report_date,
            "total":         len(all_rows),
            "total_closed":  len(closed),
            "total_pending": len(pending),
            "closed":        closed,
            "pending":       pending,
            "by_customer":   summary,
        })
    finally:
        cur.close(); conn.close()

# ── Error handlers ─────────────────────────────────────────
@app.errorhandler(404)
def not_found(_):          return err("Endpoint not found", 404)
@app.errorhandler(405)
def method_not_allowed(_): return err("Method not allowed", 405)
@app.errorhandler(500)
def internal(_):           return err("Internal server error", 500)

def seed_senthil():
    conn = get_conn(); cur = conn.cursor(dictionary=True)
    try:
        # Ensure can_delete column exists
        cur.execute("SHOW COLUMNS FROM users LIKE 'can_delete'")
        if not cur.fetchone():
            cur.execute("ALTER TABLE users ADD COLUMN can_delete TINYINT(1) NOT NULL DEFAULT 0 AFTER role")
            conn.commit()
            print("  Added can_delete column to users table")
        cur.execute("SELECT id FROM users WHERE username = %s", ("senthil",))
        existing = cur.fetchone()
        ph = bcrypt.hashpw("Senthil@007".encode(), bcrypt.gensalt()).decode()
        if existing:
            cur.execute("UPDATE users SET password_hash=%s, can_delete=1, is_active=1 WHERE username='senthil'", (ph,))
            print("  Updated senthil user — can_delete enabled")
        else:
            cur.execute(
                "INSERT INTO users (username, password_hash, full_name, role, can_delete) VALUES (%s,%s,%s,%s,%s)",
                ("senthil", ph, "Senthil", "admin", 1)
            )
            print("  Created user: senthil / Senthil@007  [can_delete=YES]")
        conn.commit()
    finally:
        cur.close(); conn.close()

# ── Entry point ────────────────────────────────────────────
if __name__ == "__main__":
    if "--create-users" in sys.argv:
        print("Seeding default users…")
        seed_users()
        print("Done.")
        sys.exit(0)
    if "--create-senthil" in sys.argv:
        print("Creating senthil user with delete privileges…")
        seed_senthil()
        print("Done.")
        sys.exit(0)
    app.run(debug=False, host="0.0.0.0", port=5000)
